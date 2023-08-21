# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from EC.models import Courses, Registro, Capacitaciones, Conferencias, Congresos, Diplomados, Talleres
from .forms import *
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic.edit import FormView
from django.urls import reverse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('administracion')
        else:
            # Si el usuario no existe o las credenciales son incorrectas, puedes mostrar un mensaje de error en la plantilla
            error_message = "Usuario o contraseña incorrectos"
            return render(request, 'login.html', {'error_message': error_message})
    return render(request, 'login.html')

def agregarCurso(request):
    courses = Courses.objects.all()
    if request.method == 'POST':
        form = CoursesForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('administracion'))
    else:
        form = CoursesForm()

    return render(request, 'administracion.html', {'courses': courses, 'form': form, 'errors': form.errors})

def editar_curso(request, curso_id):
    curso = get_object_or_404(Courses, pk=curso_id)

    if request.method == 'POST':
        imagen = request.FILES['imagen']
        titulo = request.POST['titulo']
        descripcion = request.POST['descripcion']

        curso.imagen = imagen
        curso.titulo = titulo
        curso.descripcion = descripcion
        curso.save()

        # Redirige al usuario a la página de administración con un mensaje de éxito
        return redirect('administracion')

    return render(request, 'editar_curso.html', {'curso': curso})

def eliminar_curso (request, curso_id):
    curso = get_object_or_404(Courses,pk=curso_id)

    if request.method == 'POST':
        registros_asociados = Registro.objects.filter(curso=curso)
        registros_asociados.delete()
        curso.delete()
        return redirect(administracion)
    return render(request,'eliminar_curso.html',{'curso':curso})

def agregarConferencia (request):

    conferencias = Conferencias.objects.all()
    if request.method == 'POST':
        form = ConferenciasForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('administracion'))
    else:
        form = ConferenciasForm()

    return render(request, 'administracion.html', {'courses': conferencias, 'form': form, 'errors': form.errors})

def editar_conferencia(request, conferencia_id):
    conferencia = get_object_or_404(Conferencias, pk=conferencia_id)

    if request.method == 'POST':
        imagen = request.FILES['imagen']
        titulo = request.POST['titulo']
        descripcion = request.POST['descripcion']

        conferencia.imagen = imagen
        conferencia.titulo = titulo
        conferencia.descripcion = descripcion
        conferencia.save()

        # Redirige al usuario a la página de administración con un mensaje de éxito
        return redirect('administracion')

    return render(request, 'editar_conferencia.html', {'conferencia': conferencia})

def eliminar_conferencia (request, conferencia_id):
    conferencia = get_object_or_404(Conferencias,pk=conferencia_id)
    if request.method == 'POST':
        registros_asociados = Registro.objects.filter(curso=conferencia)
        registros_asociados.delete()
        conferencia.delete()
        return redirect(administracion)
    return render(request,'eliminar_conferencia.html',{'conferencia':conferencia})

def agregarTaller (request):
    talleres = Talleres.objects.all()
    if request.method == 'POST':
        form = TalleresForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('administracion'))
    else:
        form = TalleresForm()

    return render(request, 'administracion.html', {'courses': talleres, 'form': form, 'errors': form.errors})

def editar_taller(request, taller_id):
    taller = get_object_or_404(Talleres, pk=taller_id)

    if request.method == 'POST':
        imagen = request.FILES['imagen']
        titulo = request.POST['titulo']
        descripcion = request.POST['descripcion']

        taller.imagen = imagen
        taller.titulo = titulo
        taller.descripcion = descripcion
        taller.save()

        # Redirige al usuario a la página de administración con un mensaje de éxito
        return redirect('administracion')

    return render(request, 'editar_conferencia.html', {'taller': taller})

def eliminar_taller (request, taller_id):
    taller = get_object_or_404(Talleres,pk=taller_id)
    if request.method == 'POST':
        registros_asociados = Registro.objects.filter(curso=taller)
        registros_asociados.delete()
        taller.delete()
        return redirect(administracion)
    return render(request,'eliminar_taller.html',{'taller':taller})

def agregarDiplomado(request):
    diplomados = Diplomados.objects.all()
    if request.method == 'POST':
        form = DiplomadosForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('administracion'))
    else:
        form = DiplomadosForm()

    return render(request, 'administracion.html', {'courses': diplomados, 'form': form, 'errors': form.errors})

def editar_diplomado(request, diplomado_id):
    diplomado = get_object_or_404(Diplomados, pk=diplomado_id)

    if request.method == 'POST':
        imagen = request.FILES['imagen']
        titulo = request.POST['titulo']
        descripcion = request.POST['descripcion']

        diplomado.imagen = imagen
        diplomado.titulo = titulo
        diplomado.descripcion = descripcion
        diplomado.save()

        # Redirige al usuario a la página de administración con un mensaje de éxito
        return redirect('administracion')

    return render(request, 'editar_diplomado.html', {'diplomado': diplomado})

def eliminar_diplomado(request, diplomado_id):
    diplomado = get_object_or_404(Diplomados, pk=diplomado_id)
    if request.method == 'POST':
        registros_asociados = Registro.objects.filter(curso=diplomado)
        registros_asociados.delete()
        diplomado.delete()
        return redirect('administracion')
    return render(request, 'eliminar_diplomado.html', {'diplomado': diplomado})

def agregarCongreso(request):
    congresos = Congresos.objects.all()
    if request.method == 'POST':
        form = CongresosForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('administracion'))
    else:
        form = CongresosForm()

    return render(request, 'administracion.html', {'courses': congresos, 'form': form, 'errors': form.errors})

def editar_congreso(request, congreso_id):
    congreso = get_object_or_404(Congresos, pk=congreso_id)

    if request.method == 'POST':
        imagen = request.FILES['imagen']
        titulo = request.POST['titulo']
        descripcion = request.POST['descripcion']

        congreso.imagen = imagen
        congreso.titulo = titulo
        congreso.descripcion = descripcion
        congreso.save()

        # Redirige al usuario a la página de administración con un mensaje de éxito
        return redirect('administracion')

    return render(request, 'editar_congreso.html', {'congreso': congreso})

def eliminar_congreso(request, congreso_id):
    congreso = get_object_or_404(Congresos, pk=congreso_id)
    if request.method == 'POST':
        registros_asociados = Registro.objects.filter(curso=congreso)
        registros_asociados.delete()
        congreso.delete()
        return redirect('administracion')
    return render(request, 'eliminar_congreso.html', {'congreso': congreso})

def agregarCapacitacion(request):
    capacitaciones = Capacitaciones.objects.all()
    if request.method == 'POST':
        form = CapacitacionesForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('administracion'))
    else:
        form = CapacitacionesForm()

    return render(request, 'administracion.html', {'courses': capacitaciones, 'form': form, 'errors': form.errors})

def editar_capacitacion(request, capacitacion_id):
    capacitacion = get_object_or_404(Capacitaciones, pk=capacitacion_id)

    if request.method == 'POST':
        imagen = request.FILES['imagen']
        titulo = request.POST['titulo']
        descripcion = request.POST['descripcion']

        capacitacion.imagen = imagen
        capacitacion.titulo = titulo
        capacitacion.descripcion = descripcion
        capacitacion.save()

        # Redirige al usuario a la página de administración con un mensaje de éxito
        return redirect('administracion')

    return render(request, 'editar_capacitacion.html', {'capacitacion': capacitacion})

def eliminar_capacitacion(request, capacitacion_id):
    capacitacion = get_object_or_404(Capacitaciones, pk=capacitacion_id)
    if request.method == 'POST':
        registros_asociados = Registro.objects.filter(curso=capacitacion)
        registros_asociados.delete()
        capacitacion.delete()
        return redirect('administracion')
    return render(request, 'eliminar_capacitacion.html', {'capacitacion': capacitacion})


@login_required
def administracion(request):
    # Tu código de la vista de administración aquí
    cursos = Courses.objects.all()
    conferencias = Conferencias.objects.all()
    talleres = Talleres.objects.all()
    diplomados = Diplomados.objects.all()
    congresos = Congresos.objects.all()
    capacitaciones = Capacitaciones.objects.all()
    return render(request, 'administracion.html', {'cursos': cursos, 'conferencias': conferencias, 'talleres': talleres, 'diplomados': diplomados, 'congresos': congresos, 'capacitaciones': capacitaciones})


class ReporteCursosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Registro.objects.filter(tipoRegistro='Cursos').order_by('curso')
        wb = Workbook()
        cont = 1
        ws = wb.active
        ws.title = "Hoja"+str(cont)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["B1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["C1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["D1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["E1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["F1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["G1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["H1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["I1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["J1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["K1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["L1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["A1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["B1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["C1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["D1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["E1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["F1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["G1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["H1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["I1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["J1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["K1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["L1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["A1"]="Nombres"
        ws["B1"]="Apellidos"
        ws["C1"]="E-Mail"
        ws["D1"]="Curso"
        ws["E1"]="Procedencia"
        ws["F1"]="Matricula"
        ws["G1"]="Institucion"
        ws["H1"]="Estado"
        ws["I1"]="Pais"
        ws["J1"]="Municipio"
        ws["K1"]="Estado Civil"
        ws["L1"]="Carrera"
        ws.column_dimensions["A"].width=20
        ws.column_dimensions["B"].width=20
        ws.column_dimensions["C"].width=30
        ws.column_dimensions["D"].width=45
        ws.column_dimensions["E"].width=20
        ws.column_dimensions["F"].width=25
        ws.column_dimensions["G"].width=30
        ws.column_dimensions["H"].width=20
        ws.column_dimensions["I"].width=20
        ws.column_dimensions["J"].width=20
        ws.column_dimensions["K"].width=20
        ws.column_dimensions["L"].width=30
        fila=2
        co=1
        AlCont = Alignment(vertical="center")
        BoCont = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        FoCont = Font(name= "Calibri", size= 16)
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.nombres
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.apellidos
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.email
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.curso
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.procedencia
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Matricula
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Institucion
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estado
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.pais
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.municipio
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estadocivil
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Carrera
            fila +=1
        fila = 2
        nombreArchivo = "Reporte_Cursos.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombreArchivo}"'
        wb.save(response)
        return response
        
class ReporteCapacitacionesExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Registro.objects.filter(tipoRegistro='Capacitaciones').order_by('curso')
        wb = Workbook()
        cont = 1
        ws = wb.active
        ws.title = "Hoja"+str(cont)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["B1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["C1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["D1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["E1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["F1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["G1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["H1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["I1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["J1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["K1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["L1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["A1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["B1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["C1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["D1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["E1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["F1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["G1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["H1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["I1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["J1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["K1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["L1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["A1"]="Nombres"
        ws["B1"]="Apellidos"
        ws["C1"]="E-Mail"
        ws["D1"]="Curso"
        ws["E1"]="Procedencia"
        ws["F1"]="Matricula"
        ws["G1"]="Institucion"
        ws["H1"]="Estado"
        ws["I1"]="Pais"
        ws["J1"]="Municipio"
        ws["K1"]="Estado Civil"
        ws["L1"]="Carrera"
        ws.column_dimensions["A"].width=20
        ws.column_dimensions["B"].width=20
        ws.column_dimensions["C"].width=30
        ws.column_dimensions["D"].width=45
        ws.column_dimensions["E"].width=20
        ws.column_dimensions["F"].width=25
        ws.column_dimensions["G"].width=30
        ws.column_dimensions["H"].width=20
        ws.column_dimensions["I"].width=20
        ws.column_dimensions["J"].width=20
        ws.column_dimensions["K"].width=20
        ws.column_dimensions["L"].width=30
        fila=2
        co=1
        AlCont = Alignment(vertical="center")
        BoCont = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        FoCont = Font(name= "Calibri", size= 16)
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.nombres
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.apellidos
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.email
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.curso
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.procedencia
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Matricula
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Institucion
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estado
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.pais
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.municipio
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estadocivil
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Carrera
            fila +=1
        fila = 2
        nombreArchivo = "Reporte_Capacitaciones.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombreArchivo}"'
        wb.save(response)
        return response

class ReporteConferenciasExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Registro.objects.filter(tipoRegistro='Conferencias').order_by('curso')
        wb = Workbook()
        cont = 1
        ws = wb.active
        ws.title = "Hoja"+str(cont)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["B1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["C1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["D1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["E1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["F1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["G1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["H1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["I1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["J1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["K1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["L1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["A1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["B1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["C1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["D1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["E1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["F1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["G1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["H1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["I1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["J1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["K1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["L1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["A1"]="Nombres"
        ws["B1"]="Apellidos"
        ws["C1"]="E-Mail"
        ws["D1"]="Curso"
        ws["E1"]="Procedencia"
        ws["F1"]="Matricula"
        ws["G1"]="Institucion"
        ws["H1"]="Estado"
        ws["I1"]="Pais"
        ws["J1"]="Municipio"
        ws["K1"]="Estado Civil"
        ws["L1"]="Carrera"
        ws.column_dimensions["A"].width=20
        ws.column_dimensions["B"].width=20
        ws.column_dimensions["C"].width=30
        ws.column_dimensions["D"].width=45
        ws.column_dimensions["E"].width=20
        ws.column_dimensions["F"].width=25
        ws.column_dimensions["G"].width=30
        ws.column_dimensions["H"].width=20
        ws.column_dimensions["I"].width=20
        ws.column_dimensions["J"].width=20
        ws.column_dimensions["K"].width=20
        ws.column_dimensions["L"].width=30
        fila=2
        co=1
        AlCont = Alignment(vertical="center")
        BoCont = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        FoCont = Font(name= "Calibri", size= 16)
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.nombres
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.apellidos
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.email
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.curso
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.procedencia
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Matricula
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Institucion
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estado
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.pais
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.municipio
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estadocivil
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Carrera
            fila +=1
        fila = 2
        nombreArchivo = "Reporte_Conferencias.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombreArchivo}"'
        wb.save(response)
        return response

class ReporteCongresosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Registro.objects.filter(tipoRegistro='Congresos').order_by('curso')
        wb = Workbook()
        cont = 1
        ws = wb.active
        ws.title = "Hoja"+str(cont)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["B1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["C1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["D1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["E1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["F1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["G1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["H1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["I1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["J1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["K1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["L1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["A1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["B1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["C1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["D1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["E1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["F1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["G1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["H1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["I1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["J1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["K1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["L1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["A1"]="Nombres"
        ws["B1"]="Apellidos"
        ws["C1"]="E-Mail"
        ws["D1"]="Curso"
        ws["E1"]="Procedencia"
        ws["F1"]="Matricula"
        ws["G1"]="Institucion"
        ws["H1"]="Estado"
        ws["I1"]="Pais"
        ws["J1"]="Municipio"
        ws["K1"]="Estado Civil"
        ws["L1"]="Carrera"
        ws.column_dimensions["A"].width=20
        ws.column_dimensions["B"].width=20
        ws.column_dimensions["C"].width=30
        ws.column_dimensions["D"].width=45
        ws.column_dimensions["E"].width=20
        ws.column_dimensions["F"].width=25
        ws.column_dimensions["G"].width=30
        ws.column_dimensions["H"].width=20
        ws.column_dimensions["I"].width=20
        ws.column_dimensions["J"].width=20
        ws.column_dimensions["K"].width=20
        ws.column_dimensions["L"].width=30
        fila=2
        co=1
        AlCont = Alignment(vertical="center")
        BoCont = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        FoCont = Font(name= "Calibri", size= 16)
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.nombres
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.apellidos
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.email
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.curso
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.procedencia
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Matricula
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Institucion
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estado
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.pais
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.municipio
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estadocivil
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Carrera
            fila +=1
        fila = 2
        nombreArchivo = "Reporte_Congresos.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombreArchivo}"'
        wb.save(response)
        return response

class ReporteDiplomadosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Registro.objects.filter(tipoRegistro='Diplomados').order_by('curso')
        wb = Workbook()
        cont = 1
        ws = wb.active
        ws.title = "Hoja"+str(cont)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["B1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["C1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["D1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["E1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["F1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["G1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["H1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["I1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["J1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["K1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["L1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["A1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["B1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["C1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["D1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["E1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["F1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["G1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["H1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["I1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["J1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["K1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["L1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["A1"]="Nombres"
        ws["B1"]="Apellidos"
        ws["C1"]="E-Mail"
        ws["D1"]="Curso"
        ws["E1"]="Procedencia"
        ws["F1"]="Matricula"
        ws["G1"]="Institucion"
        ws["H1"]="Estado"
        ws["I1"]="Pais"
        ws["J1"]="Municipio"
        ws["K1"]="Estado Civil"
        ws["L1"]="Carrera"
        ws.column_dimensions["A"].width=20
        ws.column_dimensions["B"].width=20
        ws.column_dimensions["C"].width=30
        ws.column_dimensions["D"].width=45
        ws.column_dimensions["E"].width=20
        ws.column_dimensions["F"].width=25
        ws.column_dimensions["G"].width=30
        ws.column_dimensions["H"].width=20
        ws.column_dimensions["I"].width=20
        ws.column_dimensions["J"].width=20
        ws.column_dimensions["K"].width=20
        ws.column_dimensions["L"].width=30
        fila=2
        co=1
        AlCont = Alignment(vertical="center")
        BoCont = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        FoCont = Font(name= "Calibri", size= 16)
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.nombres
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.apellidos
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.email
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.curso
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.procedencia
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Matricula
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Institucion
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estado
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.pais
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.municipio
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estadocivil
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Carrera
            fila +=1
        fila = 2
        nombreArchivo = "Reporte_Diplomados.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombreArchivo}"'
        wb.save(response)
        return response

class ReporteTalleresExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Registro.objects.filter(tipoRegistro='Talleres').order_by('curso')
        wb = Workbook()
        cont = 1
        ws = wb.active
        ws.title = "Hoja"+str(cont)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["B1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["C1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["D1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["E1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["F1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["G1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["H1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["I1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["J1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["K1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["L1"].border = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        ws["A1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["B1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["C1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["D1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["E1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["F1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["G1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["H1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["I1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["J1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["K1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["L1"].font = Font(name= "Calibri", size= 16 , bold=True)
        ws["A1"]="Nombres"
        ws["B1"]="Apellidos"
        ws["C1"]="E-Mail"
        ws["D1"]="Curso"
        ws["E1"]="Procedencia"
        ws["F1"]="Matricula"
        ws["G1"]="Institucion"
        ws["H1"]="Estado"
        ws["I1"]="Pais"
        ws["J1"]="Municipio"
        ws["K1"]="Estado Civil"
        ws["L1"]="Carrera"
        ws.column_dimensions["A"].width=20
        ws.column_dimensions["B"].width=20
        ws.column_dimensions["C"].width=30
        ws.column_dimensions["D"].width=45
        ws.column_dimensions["E"].width=20
        ws.column_dimensions["F"].width=25
        ws.column_dimensions["G"].width=30
        ws.column_dimensions["H"].width=20
        ws.column_dimensions["I"].width=20
        ws.column_dimensions["J"].width=20
        ws.column_dimensions["K"].width=20
        ws.column_dimensions["L"].width=30
        fila=2
        co=1
        AlCont = Alignment(vertical="center")
        BoCont = Border(left=Side(border_style= "thin"), right=Side(border_style= "thin"), bottom=Side(border_style= "thin"), top=Side(border_style= "thin"))
        FoCont = Font(name= "Calibri", size= 16)
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.nombres
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.apellidos
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.email
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.curso
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.procedencia
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Matricula
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Institucion
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estado
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.pais
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.municipio
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.estadocivil
            fila +=1
        fila = 2
        co +=1
        for q in query:
            ws.cell(row=fila, column=co).alignment = AlCont
            ws.cell(row=fila, column=co).border = BoCont
            ws.cell(row=fila, column=co).font = FoCont 
            ws.cell(row=fila, column=co).value = q.Carrera
            fila +=1
        fila = 2
        nombreArchivo = "Reporte_Talleres.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombreArchivo}"'
        wb.save(response)
        return response






